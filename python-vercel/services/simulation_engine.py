"""
Core simulation engine.
All processing is in-memory — no disk I/O except the input file bytes.
"""
from __future__ import annotations

import base64
import json
import re
from dataclasses import dataclass, field
from datetime import timedelta
from io import BytesIO
from typing import Any

import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ─── Result types ────────────────────────────────────────────────────────────

@dataclass
class TierResult:
    tier_type: str          # "basket_value" | "units"
    tier: str               # e.g. "25+ NIS" or "2+ units"
    uplift_rate: float
    forecasted_redemptions: int
    delivery_budget: float
    fixed_operational_cost: float
    final_activity_cost: float


@dataclass
class BrandSummary:
    brand: str
    quantity: float
    gmv: float


@dataclass
class SimulationOutput:
    tier_results: list[TierResult]
    brand_summary: list[BrandSummary]
    email_text: str
    xlsx_base64: str
    date_range_used: dict[str, str]
    total_orders_analyzed: int


# ─── Engine ───────────────────────────────────────────────────────────────────

class SimulationEngine:
    def __init__(self, file_bytes: bytes, config: dict[str, Any]):
        self.file_bytes = file_bytes
        self.config = config

    # ── Column resolution ────────────────────────────────────────────────────

    def _resolve_col(self, df: pd.DataFrame, col_ref: str, label: str) -> str:
        """Accept either a column letter (A, B, L…) or an exact header name."""
        if not col_ref or not col_ref.strip():
            raise HTTPException(422, detail=f"Missing column reference for: {label}")

        col_ref = col_ref.strip()

        # If it looks like a column letter (single or double letter, e.g. A, AA, BC)
        if re.match(r"^[A-Za-z]{1,3}$", col_ref) and col_ref.upper() not in df.columns:
            idx = self._letter_to_index(col_ref.upper())
            if idx >= len(df.columns):
                raise HTTPException(
                    422,
                    detail=(
                        f"Column '{col_ref}' is out of range. "
                        f"File has {len(df.columns)} columns. "
                        f"Available headers: {list(df.columns[:10])}"
                    ),
                )
            return df.columns[idx]

        # Exact header match
        if col_ref in df.columns:
            return col_ref

        # Case-insensitive match
        lower_map = {c.lower(): c for c in df.columns}
        if col_ref.lower() in lower_map:
            return lower_map[col_ref.lower()]

        raise HTTPException(
            422,
            detail=(
                f"Column '{col_ref}' not found. "
                f"Available columns: {list(df.columns[:20])}"
            ),
        )

    @staticmethod
    def _letter_to_index(letter: str) -> int:
        """Convert Excel column letter (A, B, …, Z, AA, AB, …) to 0-based index."""
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1

    # ── Load & filter ─────────────────────────────────────────────────────────

    def _load_and_filter(self) -> tuple[pd.DataFrame, dict[str, str]]:
        config = self.config

        try:
            df = pd.read_excel(BytesIO(self.file_bytes))
        except Exception as e:
            raise HTTPException(422, detail=f"Cannot read Excel file: {e}")

        if df.empty:
            raise HTTPException(422, detail="The uploaded file is empty.")

        # Resolve column names
        date_col = self._resolve_col(df, config["date_col"], "date_col")
        order_id_col = self._resolve_col(df, config["order_id_col"], "order_id_col")
        basket_value_col = self._resolve_col(df, config["basket_value_col"], "basket_value_col")
        quantity_col = self._resolve_col(df, config["quantity_col"], "quantity_col")
        gmv_col = self._resolve_col(df, config["gmv_col"], "gmv_col")
        brand_col = self._resolve_col(df, config["brand_col"], "brand_col")

        # Store resolved names back into config
        config["_date_col"] = date_col
        config["_order_id_col"] = order_id_col
        config["_basket_value_col"] = basket_value_col
        config["_quantity_col"] = quantity_col
        config["_gmv_col"] = gmv_col
        config["_brand_col"] = brand_col

        # Parse dates
        try:
            df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        except Exception:
            raise HTTPException(422, detail=f"Cannot parse dates in column '{date_col}'.")

        valid_dates = df[date_col].dropna()
        if valid_dates.empty:
            raise HTTPException(422, detail=f"No valid dates found in column '{date_col}'.")

        max_date = valid_dates.max()
        lookback = int(config.get("lookback_days", 14))
        cutoff = max_date - timedelta(days=lookback)

        df_recent = df[df[date_col] >= cutoff].copy()

        if df_recent.empty:
            raise HTTPException(
                422,
                detail=(
                    f"No orders found in the last {lookback}-day window "
                    f"(cutoff: {cutoff.date()}, latest date in file: {max_date.date()})."
                ),
            )

        date_range = {
            "from": cutoff.strftime("%d/%m/%Y"),
            "to": max_date.strftime("%d/%m/%Y"),
        }

        return df_recent, date_range

    # ── Basket value tiers ───────────────────────────────────────────────────

    def _compute_basket_value_tiers(
        self,
        df: pd.DataFrame,
        tiers: list[float],
        uplift_rates: list[float],
        delivery_cost: float,
        fixed_cost: float,
    ) -> list[TierResult]:
        results = []
        basket_col = self.config["_basket_value_col"]
        order_col = self.config["_order_id_col"]

        try:
            df[basket_col] = pd.to_numeric(df[basket_col], errors="coerce")
        except Exception:
            pass

        for threshold in sorted(tiers):
            eligible = df[df[basket_col] >= threshold]
            baseline_count = eligible[order_col].nunique()

            for rate in sorted(uplift_rates):
                forecasted = round(baseline_count * (1 + rate))
                budget = round(forecasted * delivery_cost, 2)
                total = round(fixed_cost + budget, 2)
                results.append(
                    TierResult(
                        tier_type="basket_value",
                        tier=f"{threshold}+ NIS",
                        uplift_rate=rate,
                        forecasted_redemptions=forecasted,
                        delivery_budget=budget,
                        fixed_operational_cost=fixed_cost,
                        final_activity_cost=total,
                    )
                )

        return results

    # ── Unit quantity tiers ──────────────────────────────────────────────────

    def _compute_unit_tiers(
        self,
        df: pd.DataFrame,
        tiers: list[float],
        uplift_rates: list[float],
        delivery_cost: float,
        fixed_cost: float,
    ) -> list[TierResult]:
        results = []
        qty_col = self.config["_quantity_col"]
        order_col = self.config["_order_id_col"]

        try:
            df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
        except Exception:
            pass

        order_units: pd.Series = df.groupby(order_col)[qty_col].sum()

        for threshold in sorted(tiers):
            baseline_count = int((order_units >= threshold).sum())

            for rate in sorted(uplift_rates):
                forecasted = round(baseline_count * (1 + rate))
                budget = round(forecasted * delivery_cost, 2)
                total = round(fixed_cost + budget, 2)
                results.append(
                    TierResult(
                        tier_type="units",
                        tier=f"{int(threshold)}+ units",
                        uplift_rate=rate,
                        forecasted_redemptions=forecasted,
                        delivery_budget=budget,
                        fixed_operational_cost=fixed_cost,
                        final_activity_cost=total,
                    )
                )

        return results

    # ── Brand summary ────────────────────────────────────────────────────────

    def _compute_brand_summary(self, df: pd.DataFrame) -> list[BrandSummary]:
        brand_col = self.config["_brand_col"]
        qty_col = self.config["_quantity_col"]
        gmv_col = self.config["_gmv_col"]

        try:
            df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
            df[gmv_col] = pd.to_numeric(df[gmv_col], errors="coerce").fillna(0)
        except Exception:
            pass

        agg = (
            df.groupby(brand_col)
            .agg(quantity=(qty_col, "sum"), gmv=(gmv_col, "sum"))
            .reset_index()
            .sort_values("gmv", ascending=False)
        )

        return [
            BrandSummary(brand=str(row[brand_col]), quantity=row["quantity"], gmv=row["gmv"])
            for _, row in agg.iterrows()
        ]

    # ── Output XLSX ──────────────────────────────────────────────────────────

    def _build_xlsx(
        self,
        tier_results: list[TierResult],
        brand_summary: list[BrandSummary],
    ) -> str:
        wb = Workbook()

        # ── Sheet 1: Tier Comparison ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Tier Comparison"

        yellow_fill = PatternFill("solid", fgColor="FFCC00")
        header_font = Font(bold=True, size=10)

        headers = [
            "Tier Type", "Tier", "Uplift %",
            "Forecasted Redemptions", "Delivery Budget (NIS)",
            "Fixed Operational Cost (NIS)", "Final Activity Cost (NIS)",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.fill = yellow_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(tier_results, 2):
            ws1.cell(row=row_idx, column=1, value="By Value" if r.tier_type == "basket_value" else "By Units")
            ws1.cell(row=row_idx, column=2, value=r.tier)
            ws1.cell(row=row_idx, column=3, value=f"+{int(r.uplift_rate * 100)}%")
            ws1.cell(row=row_idx, column=4, value=r.forecasted_redemptions)
            ws1.cell(row=row_idx, column=5, value=r.delivery_budget)
            ws1.cell(row=row_idx, column=6, value=r.fixed_operational_cost)
            ws1.cell(row=row_idx, column=7, value=r.final_activity_cost)

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

        # ── Sheet 2: Brand Summary ───────────────────────────────────────────
        ws2 = wb.create_sheet("Brand Summary")
        brand_headers = ["Brand / Product Type", "Quantity (Units)", "GMV (NIS)"]
        for col_idx, h in enumerate(brand_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = yellow_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, b in enumerate(brand_summary, 2):
            ws2.cell(row=row_idx, column=1, value=b.brand)
            ws2.cell(row=row_idx, column=2, value=round(b.quantity))
            ws2.cell(row=row_idx, column=3, value=round(b.gmv, 2))

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("utf-8")

    # ── Main entry point ─────────────────────────────────────────────────────

    def run(self) -> SimulationOutput:
        try:
            from app.services.email_generator import generate_hebrew_email
        except ImportError:
            from services.email_generator import generate_hebrew_email

        config = self.config
        df_recent, date_range = self._load_and_filter()

        uplift_rates = config.get("uplift_rates", [0.3, 0.5])
        basket_tiers = config.get("basket_tiers_value", [25, 30, 35])
        unit_tiers = config.get("basket_tiers_units", [2, 3])
        delivery_cost = float(config.get("delivery_cost_per_order", 15))
        fixed_cost = float(config.get("fixed_operational_cost", 5000))

        tier_results = (
            self._compute_basket_value_tiers(df_recent, basket_tiers, uplift_rates, delivery_cost, fixed_cost)
            + self._compute_unit_tiers(df_recent, unit_tiers, uplift_rates, delivery_cost, fixed_cost)
        )

        brand_summary = self._compute_brand_summary(df_recent)

        total_orders = df_recent[config["_order_id_col"]].nunique()

        xlsx_b64 = self._build_xlsx(tier_results, brand_summary)

        email_text = generate_hebrew_email(
            tier_results=tier_results,
            brand_summary=brand_summary,
            supplier_name=config.get("supplier_name", ""),
            contact_name=config.get("contact_name", ""),
            date_range=date_range,
        )

        return SimulationOutput(
            tier_results=tier_results,
            brand_summary=brand_summary,
            email_text=email_text,
            xlsx_base64=xlsx_b64,
            date_range_used=date_range,
            total_orders_analyzed=total_orders,
        )
